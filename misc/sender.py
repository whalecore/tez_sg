import smtplib
import ssl
import time
from string import Template
from openpyxl import load_workbook
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime


class ExcelRead(object):

    def __init__(self, file):
        self.file = file

    def create_book(self):
        wb = load_workbook(filename=self.file)
        sheet = wb.active
        bookings = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            agency = row[7]
            email = row[9]
            bookings[agency] = {'email': email, 'bookings': []}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            for k, v in bookings.items():
                if row[7] == k:
                    bookings[k]['bookings'].append(str(row[0]))

        return bookings

    def read_template(self, filename):
        with open(filename, 'r', encoding='utf-8') as template_file:
            template_file_content = template_file.read()
        return Template(template_file_content)


class Sender(object):

    host = 'smtp.tez-tour.com'
    port = 465
    context = ssl.create_default_context()
    server = smtplib.SMTP_SSL(host, port, context=context)

    def __init__(self, username, password):
        self.username = username
        self.password = password

    def __login(self):
        self.server.login(self.username, self.password)

    def __quit(self):
        self.server.quit()

    def send(self, filename, template):
        exread = ExcelRead(filename)
        bookings = exread.create_book()
        message_template = exread.read_template(template)
        self.__login()
        for k, v in bookings.items():
            msg = MIMEMultipart()
            bknms = ', '.join(v['bookings'])
            message = message_template.substitute(BOOKNUM=bknms)
            # print(message)
            msg['Subject'] = 'Важная информация ⚠ Перебронирование туров в Турцию'
            msg['From'] = 'kazan@tez-tour.com'
            msg['To'] = v['email']
            msg.attach(MIMEText(message, 'html'))
            self.server.send_message(msg)
            logname = datetime.now().strftime('%d%m%Y')
            with open(f'logs/{logname}.txt', 'a') as log:
                log.write(
                    f'{datetime.now().strftime("%d/%m/%Y, %H:%M")} sent to: {k}, email: {v["email"]}, booknums: {v["bookings"]}\n')
            print(f'Sending to {v["email"]}')
            time.sleep(5)
        self.__quit()

# sender = Sender()
# sender.send('excels/turkeyapr_may.xlsx', 'info_template.html')
